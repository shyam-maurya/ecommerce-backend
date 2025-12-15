from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from apps.products.models import Product
from django.db import transaction


class ProductService:
    @staticmethod
    def create_or_update(name, description, price, stock):
        product, created = Product.objects.update_or_create(
            name=name,
            defaults={
                "description": description,
                "price": price,
                "stock_quantity": stock
            }
        )
        return created  # True = created, False = updated


class Command(BaseCommand):
    help = "Import products from an Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")

    def handle(self, *args, **kwargs):
        file_path = kwargs["file_path"]

        try:
            wb = load_workbook(filename=file_path)
            ws = wb.active
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error opening file: {e}"))
            return

        created = 0
        updated = 0
        failed = 0

        # Skip header row → start from row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                name, description, price, stock = row

                if not name or price is None or stock is None:
                    failed += 1
                    continue

                with transaction.atomic():
                    is_created = ProductService.create_or_update(
                        name=name,
                        description=description,
                        price=price,
                        stock=stock
                    )

                if is_created:
                    created += 1
                else:
                    updated += 1

            except Exception as e:
                failed += 1
                self.stdout.write(self.style.ERROR(f"Error processing row {row}: {e}"))

        self.stdout.write(self.style.SUCCESS(f"Products created: {created}"))
        self.stdout.write(self.style.SUCCESS(f"Products updated: {updated}"))
        self.stdout.write(self.style.WARNING(f"Products failed: {failed}"))
